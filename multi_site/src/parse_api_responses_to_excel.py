import openpyxl as xl


def dict_to_worksheet(d, wb, row_idx, description, name):
    """
    Write dictionary of scalar results to worksheet. Creates the worksheet using `name` as the name if it does not
    exist.
    :param d: dictionary from API JSON response
    :param wb: openpyxl excel workbook
    :param row_idx: equivalent to scenario index; the row in the worksheet to write results
    :param description: scenario description; used as the left-most entry in the row
    :param name: worksheet name
    :return: nested dictionary with time-series values for each API "object";
        eg. {'PV':
                {'year_one_power_production_series_kw : [ 0.0, ..., 0.0] }
            }
    """
    time_series_dict = dict()  # for collecting lists
    if row_idx == 0:
        ws = wb.create_sheet(name)
    else:
        ws = wb[name]

    keys = [k for k in d.keys() if k.islower()]
    ws.cell(row=row_idx + 2, column=1, value=description)

    col_idx = 0
    for hdr in sorted(keys):
        if isinstance(d[hdr], list):
            list_of_lists = d[hdr]
            if any(isinstance(el, list) for el in list_of_lists):
                count = 1
                for i in list_of_lists:
                    time_series_dict[hdr + '_' + str(count)] = i
                    count = count+1
            else:
                time_series_dict[hdr] = d[hdr]
            continue
        if isinstance(d[hdr], dict) or 'series' in hdr.lower():
            continue
        if row_idx == 0:  # write header
            ws.cell(row=row_idx + 1, column=col_idx + 2, value=hdr)
        val = d[hdr]
        ws.cell(row=row_idx + 2, column=col_idx + 2, value=val)
        col_idx += 1

    return time_series_dict


def time_series_to_worksheet(d, wb, site_id):
    """
    Writes out all time series values from `d` to a worksheet in `wb` with the name: site_id + '_time_series'
    :param d: nested dictionary of time series (from dict_to_worksheet)
    :param wb: openpyxl excel workbook
    :param description: scenario description; used to name worksheet: site_id + '_time_series'
    :return: None
    """
    ws = wb.create_sheet(site_id + '_time_series')
    col_idx = 0
    for K in [Key for Key in d.keys() if len(d[Key].keys()) > 0]:
        for k in [key for key in d[K].keys() if len(d[K][key]) > 0]:
            ws.cell(row=1, column=col_idx + 1, value=K + '|' + k)
            for j, val in enumerate(d[K][k]):
                ws.cell(row=2 + j, column=col_idx + 1, value=val)
            col_idx += 1


def parse_api_responses_to_excel(responses, spreadsheet='results_summary.xlsx'):
    """
    Writes all inputs, outputs, and dispatches to separate worksheets in an Excel workbook.
    REopt object inputs and outputs are on separate sheets (eg. 'PV_inputs' and 'PV_outputs') with each row
    corresponding to a scenario.
    Dispatch time-series for each scenario are written to one worksheet per scenario.
    :param responses: list of dictionaries - one dict = one api response
    :param spreadsheet: str, path/to/excel_spreadheet.xlsx to write results summary
    :return: None
    """
    wb = xl.load_workbook('inputs/template.xlsx')
    ws_sites = wb.active

    for row_idx, resp in enumerate(responses):
        time_series_dict = dict()
        inputs_dict = resp['inputs']['Scenario']
        outputs_dict = resp['outputs']['Scenario']
        description = inputs_dict['description']
        site_id = inputs_dict['site_id']
        ws_sites.cell(row=row_idx + 2, column=1, value=site_id)
        scenario_inputs = [k for k in inputs_dict.keys() if k.islower()]
        site_inputs = [k for k in inputs_dict['Site'].keys() if k.islower()]

        for col_idx, hdr in enumerate(sorted(scenario_inputs) + sorted(site_inputs)):
#             if hdr == 'site_id':
#                 continue
            if row_idx == 0:  # write header
                ws_sites.cell(row=row_idx + 1, column=col_idx + 1, value=hdr)
            try:
                val = inputs_dict[hdr]
            except KeyError:
                val = inputs_dict['Site'][hdr]
            ws_sites.cell(row=row_idx + 2, column=col_idx + 1, value=val)

        err_msg = resp['messages'].get('error')
        if row_idx == 0:  # write header
            ws_sites.cell(row=row_idx + 1, column=col_idx + 2, value='error_message')
        ws_sites.cell(row=row_idx + 2, column=col_idx + 2, value=err_msg)

        for Key in [K for K in inputs_dict['Site'].keys() if K[0].isupper()]:
            time_series_dict[Key] = dict_to_worksheet(inputs_dict['Site'][Key], wb, row_idx, description, Key + '_inputs')

        for Key in [K for K in outputs_dict['Site'].keys() if K[0].isupper()]:
            time_series_dict[Key] = dict_to_worksheet(outputs_dict['Site'][Key], wb, row_idx, description, Key + '_outputs')

#         time_series_to_worksheet(time_series_dict, wb, str(site_id))
    wb.save(spreadsheet)


if __name__ == '__main__':  # test script
    import json
    import os
    resp1 = json.load(open(os.path.join('..', 'outputs', 'site3-No-PV.json'), 'rb'))
    resp2 = json.load(open(os.path.join('..', 'outputs', 'site3-With-PV.json'), 'rb'))

    responses = [resp1, resp2]
    parse_api_responses_to_excel(responses, 'test.xlsx')
